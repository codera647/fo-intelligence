"""Validation module — verifies enriched records and exports final dataset.

Exports XLSX matching example dataset formatting:
  - Verdana 12pt bold centered headers with tier-based coloring
  - Garamond 11pt data font, single-line (no wrap), vertical=top
  - Compact row height (15pt data rows)
  - Medium borders, conditional formatting on scores
  - Auto-filters, frozen panes, proper column widths
"""

import logging
import pandas as pd
from typing import List, Dict
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter

from config.schema import COLUMN_ORDER, FamilyOfficeRecord
from config.settings import PROCESSED_DIR

logger = logging.getLogger(__name__)

# ── Color palette ─────────────────────────────────────────────────
TIER1_HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
TIER2_HEADER_FILL = PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid")
TIER3_HEADER_FILL = PatternFill(start_color="3D7AB5", end_color="3D7AB5", fill_type="solid")
TIER4_HEADER_FILL = PatternFill(start_color="4A8CC7", end_color="4A8CC7", fill_type="solid")

HEADER_FONT = Font(name="Verdana", size=12, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Garamond", size=11, color="333333")
NAME_FONT = Font(name="Garamond", size=11, bold=True, color="1B3A5C")

MEDIUM_BORDER = Border(
    left=Side(style="medium", color="B4C6E0"),
    right=Side(style="medium", color="B4C6E0"),
    top=Side(style="medium", color="B4C6E0"),
    bottom=Side(style="medium", color="B4C6E0"),
)
HEADER_BORDER = Border(
    left=Side(style="medium", color="0D2137"),
    right=Side(style="medium", color="0D2137"),
    top=Side(style="medium", color="0D2137"),
    bottom=Side(style="thick", color="0D2137"),
)

ROW_FILL_EVEN = PatternFill(start_color="EAF0F7", end_color="EAF0F7", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ENTITY_COLORS = {
    "Single Family Office": PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),
    "Multi Family Office": PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"),
    "Hybrid": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
}

URL_QUALITY_FILLS = {
    "Highest": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Medium-Low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Lower": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Not Found": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
URL_QUALITY_FONTS = {
    "Highest": Font(name="Garamond", size=11, color="006100"),
    "Medium": Font(name="Garamond", size=11, color="9C6500"),
    "Medium-Low": Font(name="Garamond", size=11, color="9C5700"),
    "Lower": Font(name="Garamond", size=11, color="C00000"),
    "Not Found": Font(name="Garamond", size=11, color="808080"),
}

TIER1_COLS = {
    "family_office_name", "entity_type", "description", "year_founded",
    "aum_estimated", "aum_source", "investment_thesis", "investing_sectors",
    "website_url", "url_quality", "corporate_linkedin_url",
    "corporate_email", "other_socials",
    "hq_city", "hq_state", "hq_country",
}
TIER2_COLS = {
    "contact_name", "contact_title", "contact_linkedin", "contact_email",
    "email_confidence", "email_source",
}
TIER3_COLS = {
    "recent_activity", "activity_date", "activity_source_url", "key_investments",
}
TIER4_COLS = {
    "data_completeness_score", "confidence_score",
}

COL_WIDTHS = {
    "family_office_name": 28, "entity_type": 20, "description": 45,
    "year_founded": 13, "aum_estimated": 20, "aum_source": 18,
    "investment_thesis": 42, "investing_sectors": 28, "website_url": 32,
    "url_quality": 14, "corporate_linkedin_url": 32,
    "corporate_email": 28, "other_socials": 40,
    "hq_city": 16,
    "hq_state": 16, "hq_country": 22, "contact_name": 22,
    "contact_title": 22, "contact_linkedin": 32, "contact_email": 26,
    "email_confidence": 16, "email_source": 16, "recent_activity": 40,
    "activity_date": 14, "activity_source_url": 32, "key_investments": 40,
    "data_completeness_score": 18, "confidence_score": 16,
}

HEADER_LABELS = {
    "family_office_name": "Family Office Name", "entity_type": "Entity Type",
    "description": "Description", "year_founded": "Year Founded",
    "aum_estimated": "AUM (Estimated)", "aum_source": "AUM Source",
    "investment_thesis": "Investment Thesis", "investing_sectors": "Investing Sectors",
    "website_url": "Website URL", "url_quality": "URL Quality",
    "corporate_linkedin_url": "Corporate LinkedIn",
    "corporate_email": "Corporate Email", "other_socials": "Other Socials",
    "hq_city": "HQ City",
    "hq_state": "HQ State / Region", "hq_country": "HQ Country",
    "contact_name": "Contact Name", "contact_title": "Contact Title",
    "contact_linkedin": "Contact LinkedIn", "contact_email": "Contact Email",
    "email_confidence": "Email Confidence", "email_source": "Email Source",
    "recent_activity": "Recent Activity", "activity_date": "Activity Date",
    "activity_source_url": "Activity Source URL", "key_investments": "Key Investments",
    "data_completeness_score": "Completeness Score", "confidence_score": "Confidence Score",
}


def _sanitize_single_line(value) -> str:
    """Strip newlines, bullets, excess whitespace for single-line cells."""
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = s.replace("•", ",").replace("–", "-").replace("—", "-")
    s = s.replace("  ", " ").strip()
    return s


# Leading characters that spreadsheet apps (Excel/Sheets) treat as a formula.
# Crawled/LLM-derived text is untrusted, so neutralize these to prevent CSV
# formula injection (e.g. "=HYPERLINK(...)", "=cmd|'...'!A1", "@SUM(...)").
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _neutralize_formula(value):
    """Prefix a leading formula trigger with a single quote so the cell stays text.

    Only touches strings that begin with a dangerous character; all other
    values (numbers, None, safe strings) pass through unchanged.
    """
    if isinstance(value, str) and value and value[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _get_tier_fill(col: str) -> PatternFill:
    """Return the header fill colour for a column based on its tier."""
    if col in TIER1_COLS:
        return TIER1_HEADER_FILL
    if col in TIER2_COLS:
        return TIER2_HEADER_FILL
    if col in TIER3_COLS:
        return TIER3_HEADER_FILL
    return TIER4_HEADER_FILL


def _clean_record(rec: dict) -> dict:
    """Ensure every column exists and recalculate completeness."""
    cleaned = {}
    for col in COLUMN_ORDER:
        val = rec.get(col)
        if isinstance(val, str):
            val = _sanitize_single_line(val)
            if val.lower() in ("none", "n/a", "unknown", ""):
                val = None
        cleaned[col] = val

    # Recalculate completeness
    filled = sum(1 for c in COLUMN_ORDER if cleaned.get(c) not in (None, "", "Not Found"))
    cleaned["data_completeness_score"] = round(filled / len(COLUMN_ORDER) * 100, 1)
    return cleaned


def validate_records(records: list[dict]) -> list[dict]:
    """Clean, deduplicate, sort by confidence, return validated list."""
    seen_names = set()
    validated = []

    for rec in records:
        name = rec.get("family_office_name", "").strip()
        if not name or name.lower() in seen_names:
            continue
        seen_names.add(name.lower())

        cleaned = _clean_record(rec)
        validated.append(cleaned)

    validated.sort(key=lambda r: r.get("confidence_score") or 0, reverse=True)
    logger.info(f"Validated {len(validated)} unique records from {len(records)} input")
    return validated


def export_to_xlsx(records: list[dict], filepath: Path = None) -> Path:
    """Export records to a professionally formatted XLSX file."""
    if filepath is None:
        filepath = PROCESSED_DIR / "family_offices_dataset.xlsx"
    filepath.parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(records)
    for col in COLUMN_ORDER:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMN_ORDER]

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Family Offices", index=False)
        ws = writer.sheets["Family Offices"]

        # ── Column widths ───────────────────────────────────────────
        for i, col in enumerate(COLUMN_ORDER, 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = COL_WIDTHS.get(col, 18)

        # ── Header row (row 1) ──────────────────────────────────────
        ws.row_dimensions[1].height = 50
        for i, col in enumerate(COLUMN_ORDER, 1):
            cell = ws.cell(row=1, column=i)
            cell.value = HEADER_LABELS.get(col, col)
            cell.font = HEADER_FONT
            cell.fill = _get_tier_fill(col)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = HEADER_BORDER

        # ── Data rows ───────────────────────────────────────────────
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 15
            row_fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD

            for col_idx in range(1, len(COLUMN_ORDER) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_name = COLUMN_ORDER[col_idx - 1]

                # Sanitize value + neutralize spreadsheet formula injection.
                # (openpyxl treats a str starting with "=" as a live formula.)
                if cell.value is not None:
                    cell.value = _neutralize_formula(_sanitize_single_line(cell.value))

                # Font
                if col_name == "family_office_name":
                    cell.font = NAME_FONT
                else:
                    cell.font = DATA_FONT

                # Alignment — single line, no wrap
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=False
                )

                # Fill
                cell.fill = row_fill
                cell.border = MEDIUM_BORDER

                # ── Conditional styling ─────────────────────────────
                val = cell.value

                # Entity type coloring
                if col_name == "entity_type" and val in ENTITY_COLORS:
                    cell.fill = ENTITY_COLORS[val]

                # URL quality coloring
                if col_name == "url_quality" and val in URL_QUALITY_FILLS:
                    cell.fill = URL_QUALITY_FILLS[val]
                    cell.font = URL_QUALITY_FONTS.get(val, DATA_FONT)

                # Score formatting
                if col_name in ("data_completeness_score", "confidence_score"):
                    if val is not None:
                        try:
                            num = float(val)
                            cell.value = num
                            cell.number_format = "0.0"
                            if num >= 70:
                                cell.font = Font(name="Garamond", size=11, color="006100", bold=True)
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif num >= 40:
                                cell.font = Font(name="Garamond", size=11, color="9C6500")
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.font = Font(name="Garamond", size=11, color="C00000")
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        except (ValueError, TypeError):
                            pass

                # Email confidence coloring
                if col_name == "email_confidence":
                    conf_colors = {
                        "Verified": ("006100", "C6EFCE"),
                        "High": ("006100", "C6EFCE"),
                        "Medium": ("9C6500", "FFEB9C"),
                        "Low": ("C00000", "FFC7CE"),
                        "Not Found": ("808080", "D9D9D9"),
                    }
                    if val in conf_colors:
                        fc, bc = conf_colors[val]
                        cell.font = Font(name="Garamond", size=11, color=fc)
                        cell.fill = PatternFill(start_color=bc, end_color=bc, fill_type="solid")

        # ── Freeze panes & auto-filter ──────────────────────────────
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

    logger.info(f"Exported {len(records)} records → {filepath}")
    return filepath


def export_to_csv(records: list[dict], filepath: Path = None) -> Path:
    """Export records to CSV."""
    if filepath is None:
        filepath = PROCESSED_DIR / "family_offices_dataset.csv"
    filepath.parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(records)
    for col in COLUMN_ORDER:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMN_ORDER]

    # Sanitize all string values + neutralize spreadsheet formula injection
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: _neutralize_formula(_sanitize_single_line(v)) if isinstance(v, str) else v
        )

    df.to_csv(filepath, index=False)
    logger.info(f"Exported {len(records)} records → {filepath}")
    return filepath


def generate_stats(records: list[dict]) -> dict:
    """Generate summary statistics for the dataset."""
    total = len(records)
    if total == 0:
        return {"total": 0}

    df = pd.DataFrame(records)

    def pct_filled(col):
        if col not in df.columns:
            return 0.0
        return round(df[col].notna().sum() / total * 100, 1)

    stats = {
        "total_records": total,
        "avg_completeness": round(df["data_completeness_score"].mean(), 1) if "data_completeness_score" in df.columns else 0,
        "avg_confidence": round(df["confidence_score"].mean(), 1) if "confidence_score" in df.columns else 0,
        "field_coverage": {col: pct_filled(col) for col in COLUMN_ORDER},
        "entity_types": df["entity_type"].value_counts().to_dict() if "entity_type" in df.columns else {},
        "url_quality_dist": df["url_quality"].value_counts().to_dict() if "url_quality" in df.columns else {},
        "email_confidence_dist": df["email_confidence"].value_counts().to_dict() if "email_confidence" in df.columns else {},
        "records_with_website": int(df["website_url"].notna().sum()) if "website_url" in df.columns else 0,
        "records_with_contact": int(df["contact_name"].notna().sum()) if "contact_name" in df.columns else 0,
        "records_with_email": int(df["contact_email"].notna().sum()) if "contact_email" in df.columns else 0,
        "records_with_linkedin": int(df["contact_linkedin"].notna().sum()) if "contact_linkedin" in df.columns else 0,
    }
    return stats
